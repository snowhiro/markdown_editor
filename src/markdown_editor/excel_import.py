"""Excel → Markdown 変換（spec.md 11章）。

シート単位でMarkdownを生成する純粋なロジックを提供する。
Qtに依存しないため、単体でテストできる（tests/test_excel_import.py）。

テーブル領域はセルの内容ではなく**罫線の有無**で判定する（spec.md 11.4）。
描画オブジェクト（画像・グラフ・図形）は変換対象外だが、存在する場合は
その旨をmdの末尾に注記する（spec.md 11.8）。
"""

from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Callable
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

SUPPORTED_SUFFIXES = (".xlsx", ".xlsm")

# テーブルと見なす最小サイズ（spec.md 11.4）。
# 「見出しの下に罫線を1本引いただけ」を表として誤検知しないための下限。
MIN_TABLE_ROWS = 2
MIN_TABLE_COLS = 2


# ---------------------------------------------------------------------------
# 結果オブジェクト
# ---------------------------------------------------------------------------


@dataclass
class SheetResult:
    """シート1枚の変換結果。"""

    sheet_name: str
    ok: bool
    path: Path | None = None
    error: str | None = None
    # 出力しなかった理由。現状は内容が無い場合の "empty" のみ
    # （非表示シートはそもそも処理対象に含めないため結果に現れない）
    skipped: str | None = None


@dataclass
class ConversionResult:
    out_dir: Path
    sheets: list[SheetResult] = field(default_factory=list)
    # 数式のキャッシュ値を取得できなかったシート名（spec.md 11.5）
    missing_formula_sheets: list[str] = field(default_factory=list)

    @property
    def written(self) -> list[Path]:
        return [s.path for s in self.sheets if s.ok and s.path is not None]

    @property
    def failed(self) -> list[SheetResult]:
        return [s for s in self.sheets if not s.ok]


# ---------------------------------------------------------------------------
# ファイル名のサニタイズ（spec.md 11.2）
# ---------------------------------------------------------------------------

# Excelはシート名に / \ : * ? [ ] を許可しないが、< > " | ・前後の空白・
# Windowsの予約名は許容されるため、ファイル名としては不正になり得る。
# Excel側が拒否する文字も、他ツール生成のファイルに備えて併せて置換する。
_INVALID_NAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

_WINDOWS_RESERVED = {
    "CON", "PRN", "AUX", "NUL",
    *(f"COM{i}" for i in range(1, 10)),
    *(f"LPT{i}" for i in range(1, 10)),
}


def sanitize_sheet_name(name: str) -> str:
    """シート名をファイル名（拡張子なし）として安全な文字列に正規化する。"""
    cleaned = _INVALID_NAME_CHARS.sub("_", name)
    # 末尾の空白・ピリオドはWindowsが暗黙に落とすため、先に取り除く
    cleaned = cleaned.strip().strip(".").strip()
    if not cleaned:
        return "sheet"
    if cleaned.upper() in _WINDOWS_RESERVED:
        return cleaned + "_"
    return cleaned


def unique_file_names(sheet_names: list[str]) -> list[str]:
    """シート名の一覧を、重複しないmdファイル名（拡張子なし）へ変換する。

    正規化の結果が衝突した場合は -2, -3 の連番を付ける。
    比較はWindows/macOSのファイルシステムに合わせて大文字小文字を区別しない。
    """
    used: set[str] = set()
    result: list[str] = []
    for name in sheet_names:
        base = sanitize_sheet_name(name)
        candidate = base
        counter = 2
        while candidate.casefold() in used:
            candidate = f"{base}-{counter}"
            counter += 1
        used.add(candidate.casefold())
        result.append(candidate)
    return result


# ---------------------------------------------------------------------------
# 表示形式の適用（spec.md 11.5）
# ---------------------------------------------------------------------------


def _split_format_sections(fmt: str) -> list[str]:
    """表示形式を ';' 区切りのセクションへ分割する（引用符・角括弧内は無視）。"""
    sections: list[str] = []
    buf: list[str] = []
    i, n = 0, len(fmt)
    while i < n:
        ch = fmt[i]
        if ch == '"':
            end = fmt.find('"', i + 1)
            end = n - 1 if end < 0 else end
            buf.append(fmt[i : end + 1])
            i = end + 1
        elif ch == "[":
            end = fmt.find("]", i + 1)
            end = n - 1 if end < 0 else end
            buf.append(fmt[i : end + 1])
            i = end + 1
        elif ch == "\\" and i + 1 < n:
            buf.append(fmt[i : i + 2])
            i += 2
        elif ch == ";":
            sections.append("".join(buf))
            buf = []
            i += 1
        else:
            buf.append(ch)
            i += 1
    sections.append("".join(buf))
    return sections


def _iter_format_tokens(section: str):
    """表示形式を (種別, 文字列) のトークン列に分解する。

    種別: "literal"（そのまま出力）/ "skip"（出力しない）/ "code"（書式コード）
    """
    i, n = 0, len(section)
    while i < n:
        ch = section[i]
        if ch == '"':
            end = section.find('"', i + 1)
            end = n if end < 0 else end
            yield ("literal", section[i + 1 : end])
            i = end + 1
        elif ch == "[":
            # [Red] や [$-409] のような修飾子。色・ロケール指定は出力しない
            end = section.find("]", i + 1)
            end = n - 1 if end < 0 else end
            yield ("skip", section[i : end + 1])
            i = end + 1
        elif ch == "\\" and i + 1 < n:
            yield ("literal", section[i + 1])
            i += 2
        elif ch == "_" and i + 1 < n:
            # 「次の文字の幅だけ空ける」。Markdownでは幅を持たないため無視する
            yield ("skip", section[i : i + 2])
            i += 2
        elif ch == "*" and i + 1 < n:
            # 繰り返し埋め。幅の概念がないため無視する
            yield ("skip", section[i : i + 2])
            i += 2
        elif (m := _AMPM_RE.match(section, i)) is not None:
            # AM/PM は1文字ずつではなく1トークンとして扱う
            yield ("code", m.group(0))
            i = m.end()
        else:
            yield ("code", ch)
            i += 1


def _is_date_section(section: str) -> bool:
    """セクションが日付/時刻の表示形式かを判定する。"""
    for kind, text in _iter_format_tokens(section):
        if kind == "code" and text.lower() in "ymdhs":
            return True
    return False


_AMPM_RE = re.compile(r"(AM/PM|A/P)", re.IGNORECASE)


def _to_datetime(value) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    if isinstance(value, time):
        return datetime(1899, 12, 30, value.hour, value.minute, value.second)
    return None


def _format_date_section(value, section: str) -> str:
    """日付/時刻の表示形式を適用する。"""
    dt = _to_datetime(value)
    if dt is None:
        return _default_text(value)

    # 先に code トークンを連続同種でまとめる（yyyy, mm, ss ...）
    tokens: list[tuple[str, str]] = []
    for kind, text in _iter_format_tokens(section):
        if kind == "code" and tokens and tokens[-1][0] == "code" and tokens[-1][1][0] == text:
            tokens[-1] = ("code", tokens[-1][1] + text)
        else:
            tokens.append((kind, text))

    has_ampm = bool(_AMPM_RE.search(section))
    out: list[str] = []
    i = 0
    while i < len(tokens):
        kind, text = tokens[i]
        i += 1
        if kind == "skip":
            continue
        if kind == "literal":
            out.append(text)
            continue

        if text.upper() in ("AM/PM", "A/P"):
            marker = ("AM", "PM") if text.upper() == "AM/PM" else ("A", "P")
            out.append(marker[0] if dt.hour < 12 else marker[1])
            continue

        head = text[0].lower()
        width = len(text)
        if head == "y":
            out.append(f"{dt.year:04d}" if width >= 3 else f"{dt.year % 100:02d}")
        elif head == "d":
            if width >= 4:
                out.append(_WEEKDAYS_FULL[dt.weekday()])
            elif width == 3:
                out.append(_WEEKDAYS_SHORT[dt.weekday()])
            elif width == 2:
                out.append(f"{dt.day:02d}")
            else:
                out.append(str(dt.day))
        elif head == "h":
            hour = dt.hour
            if has_ampm:
                hour = hour % 12 or 12
            out.append(f"{hour:02d}" if width >= 2 else str(hour))
        elif head == "s":
            out.append(f"{dt.second:02d}" if width >= 2 else str(dt.second))
        elif head == "m":
            # 'm' は月と分の両方を表す。直前が時、または直後が秒なら分と解釈する。
            # 区切りの ':' や '/' も code トークンなので、日時コードだけを見る。
            def nearest(seq) -> str:
                return next(
                    (t for k, t in seq if k == "code" and t[0].lower() in "ymdhs"), ""
                )

            prev_code = nearest(reversed(tokens[: i - 1]))
            next_code = nearest(tokens[i:])
            is_minute = prev_code[:1].lower() == "h" or next_code[:1].lower() == "s"
            if is_minute:
                out.append(f"{dt.minute:02d}" if width >= 2 else str(dt.minute))
            elif width >= 4:
                out.append(_MONTHS_FULL[dt.month - 1])
            elif width == 3:
                out.append(_MONTHS_SHORT[dt.month - 1])
            elif width == 2:
                out.append(f"{dt.month:02d}")
            else:
                out.append(str(dt.month))
        else:
            out.append(text)
    return "".join(out)


_MONTHS_FULL = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
_MONTHS_SHORT = [m[:3] for m in _MONTHS_FULL]
_WEEKDAYS_FULL = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
_WEEKDAYS_SHORT = [d[:3] for d in _WEEKDAYS_FULL]


def _format_number_section(value: float, section: str, keep_sign: bool) -> str:
    """数値の表示形式を適用する。"""
    literals_before: list[str] = []
    literals_after: list[str] = []
    pattern: list[str] = []
    percent = 0
    seen_pattern = False

    for kind, text in _iter_format_tokens(section):
        if kind == "skip":
            continue
        if kind == "code" and text in "0#?.,":
            pattern.append(text)
            seen_pattern = True
            continue
        if kind == "code" and text == "%":
            percent += 1
            (literals_after if seen_pattern else literals_before).append("%")
            continue
        (literals_after if seen_pattern else literals_before).append(text)

    if not pattern:
        # 数値プレースホルダを持たない書式（"完了" のような固定文字列など）
        return "".join(literals_before) + "".join(literals_after)

    # Excelの丸めは四捨五入。Pythonの既定（偶数丸め）とずれるためDecimalで揃える。
    # 2進浮動小数の誤差を持ち込まないよう repr() 経由でDecimal化する。
    try:
        number = Decimal(repr(float(value))) * (Decimal(100) ** percent)
    except (InvalidOperation, ValueError, OverflowError):
        return _default_text(value)

    body = "".join(pattern)
    if "." in body:
        int_part, frac_part = body.split(".", 1)
    else:
        int_part, frac_part = body, ""
    decimals = sum(1 for ch in frac_part if ch in "0#?")
    thousands = "," in int_part
    min_int_digits = sum(1 for ch in int_part if ch == "0")

    negative = number < 0
    rounded = abs(number).quantize(
        Decimal(1).scaleb(-decimals), rounding=ROUND_HALF_UP
    )
    text = f"{rounded:,.{decimals}f}" if thousands else f"{rounded:.{decimals}f}"
    if min_int_digits > 1:
        head, _, tail = text.partition(".")
        digits = head.replace(",", "")
        if len(digits) < min_int_digits:
            head = digits.rjust(min_int_digits, "0")
        text = head + ("." + tail if tail else "")
    if negative and keep_sign:
        text = "-" + text
    return "".join(literals_before) + text + "".join(literals_after)


def _default_text(value) -> str:
    """表示形式を適用しない既定の文字列化。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if value.is_integer() and abs(value) < 1e15:
            return str(int(value))
        return f"{value:.10g}"
    if isinstance(value, datetime):
        if (value.hour, value.minute, value.second) == (0, 0, 0):
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, timedelta):
        total = int(value.total_seconds())
        return f"{total // 3600}:{total % 3600 // 60:02d}:{total % 60:02d}"
    return str(value)


def format_value(value, number_format: str | None) -> str:
    """Excelの表示形式を適用した文字列を返す（spec.md 11.5）。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"

    fmt = (number_format or "").strip()
    if not fmt or fmt.lower() == "general" or fmt == "@":
        return _default_text(value)

    sections = _split_format_sections(fmt)
    if isinstance(value, str):
        # 4つ目のセクションは文字列用。@ を実際の文字列に差し替える
        if len(sections) >= 4:
            out: list[str] = []
            for kind, text in _iter_format_tokens(sections[3]):
                if kind == "skip":
                    continue
                out.append(value if kind == "code" and text == "@" else text)
            return "".join(out)
        return value

    if isinstance(value, (datetime, date, time)) or _is_date_section(sections[0]):
        return _format_date_section(value, sections[0])

    if isinstance(value, timedelta):
        return _default_text(value)

    # 数値: 正/負/ゼロでセクションを使い分ける
    keep_sign = True
    section = sections[0]
    if value < 0 and len(sections) >= 2 and sections[1].strip():
        section, keep_sign = sections[1], False
    elif value == 0 and len(sections) >= 3 and sections[2].strip():
        section = sections[2]
    try:
        return _format_number_section(float(value), section, keep_sign)
    except (ValueError, TypeError):
        return _default_text(value)


# ---------------------------------------------------------------------------
# セル → Markdown 断片
# ---------------------------------------------------------------------------


def decorate(text: str, *, bold: bool, italic: bool, strike: bool) -> str:
    """文字装飾をMarkdownの記法で反映する（spec.md 11.5）。

    Markdownの強調は前後に空白があると成立しないため、装飾は空白を除いた
    芯の部分にのみ適用し、元の空白は外側へ戻す。
    """
    if not text.strip() or not (bold or italic or strike):
        return text
    core = text.strip()
    lead = text[: len(text) - len(text.lstrip())]
    trail = text[len(text.rstrip()) :]
    if italic:
        core = f"*{core}*"
    if bold:
        core = f"**{core}**"
    if strike:
        core = f"~~{core}~~"
    return lead + core + trail


def cell_text(cell, *, in_table: bool) -> str:
    """セルをMarkdownの断片へ変換する（spec.md 11.5）。"""
    text = format_value(cell.value, cell.number_format)
    if not text:
        return ""

    font = cell.font
    text = decorate(
        text,
        bold=bool(font and font.bold),
        italic=bool(font and font.italic),
        strike=bool(font and font.strike),
    )

    link = getattr(cell, "hyperlink", None)
    if link is not None:
        target = link.target or ""
        if link.location:
            target = f"{target}#{link.location}" if target else f"#{link.location}"
        if target:
            text = f"[{text}]({target})"

    # セル内改行は表・段落ともに <br> で表現する
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")
    if in_table:
        text = text.replace("|", "\\|")
    return text


# ---------------------------------------------------------------------------
# テーブル判定（spec.md 11.4）
# ---------------------------------------------------------------------------

Rect = tuple[int, int, int, int]  # (row1, col1, row2, col2) いずれも1始まり・両端含む


def _has_border(cell) -> bool:
    border = cell.border
    if border is None:
        return False
    for side in (border.left, border.right, border.top, border.bottom):
        if side is not None and side.style:
            return True
    return False


def _rects_overlap(a: Rect, b: Rect) -> bool:
    return not (a[2] < b[0] or b[2] < a[0] or a[3] < b[1] or b[3] < a[1])


def _merge_overlapping(rects: list[Rect]) -> list[Rect]:
    """外接矩形が重なり合う候補どうしを1つに統合する。"""
    merged = list(rects)
    changed = True
    while changed:
        changed = False
        for i in range(len(merged)):
            for j in range(i + 1, len(merged)):
                a, b = merged[i], merged[j]
                if _rects_overlap(a, b):
                    merged[i] = (
                        min(a[0], b[0]), min(a[1], b[1]),
                        max(a[2], b[2]), max(a[3], b[3]),
                    )
                    del merged[j]
                    changed = True
                    break
            if changed:
                break
    return merged


def detect_tables(bordered: set[tuple[int, int]]) -> list[Rect]:
    """罫線セルの集合からテーブル領域（外接矩形）を求める。

    4近傍の連結成分ごとに外接矩形を取り、重なるものを統合したうえで
    2行×2列未満の候補を除外する。
    """
    remaining = set(bordered)
    rects: list[Rect] = []
    while remaining:
        start = remaining.pop()
        stack = [start]
        r1 = r2 = start[0]
        c1 = c2 = start[1]
        while stack:
            r, c = stack.pop()
            r1, r2 = min(r1, r), max(r2, r)
            c1, c2 = min(c1, c), max(c2, c)
            for nr, nc in ((r - 1, c), (r + 1, c), (r, c - 1), (r, c + 1)):
                if (nr, nc) in remaining:
                    remaining.discard((nr, nc))
                    stack.append((nr, nc))
        rects.append((r1, c1, r2, c2))

    rects = _merge_overlapping(rects)
    return sorted(
        r for r in rects
        if r[2] - r[0] + 1 >= MIN_TABLE_ROWS and r[3] - r[1] + 1 >= MIN_TABLE_COLS
    )


# ---------------------------------------------------------------------------
# 描画オブジェクトの検出（spec.md 11.8 / 13.1）
# ---------------------------------------------------------------------------

_NS_R = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_ANCHOR_TAGS = {"twoCellAnchor", "oneCellAnchor", "absoluteAnchor"}
_OBJECT_LABELS = {
    "pic": "画像",
    "sp": "図形・テキストボックス",
    "cxnSp": "コネクタ",
    "grpSp": "グループ図形",
}


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def scan_drawing_objects(xlsx_path: Path) -> dict[str, dict[str, int]]:
    """シート名 → 種別ごとの描画オブジェクト数 を返す。

    openpyxl は未対応の図形を含む描画パートを画像・グラフごと読み捨てるため
    （spec.md 13.1）、検出だけはxlsxのzipを直接読んで行う。
    """
    counts: dict[str, dict[str, int]] = {}
    try:
        with zipfile.ZipFile(xlsx_path) as zf:
            names = set(zf.namelist())
            book = ET.fromstring(zf.read("xl/workbook.xml"))
            rels = _read_rels(zf, "xl/_rels/workbook.xml.rels", "xl")
            for sheet in book.iter():
                if _local(sheet.tag) != "sheet":
                    continue
                title = sheet.get("name") or ""
                part = rels.get(sheet.get(f"{_NS_R}id", ""))
                if not part or part not in names:
                    continue
                sheet_rels = _read_rels(
                    zf,
                    f"{_posix_dir(part)}/_rels/{Path(part).name}.rels",
                    _posix_dir(part),
                )
                found: dict[str, int] = {}
                for target in sheet_rels.values():
                    if not target.startswith("xl/drawings/") or not target.endswith(".xml"):
                        continue
                    if target not in names:
                        continue
                    _count_drawing_objects(ET.fromstring(zf.read(target)), found)
                if found:
                    counts[title] = found
    except (OSError, zipfile.BadZipFile, ET.ParseError, KeyError):
        # 検出は付随情報のため、失敗しても変換自体は続行する
        return counts
    return counts


def _posix_dir(part: str) -> str:
    return part.rsplit("/", 1)[0] if "/" in part else ""


def _read_rels(zf: zipfile.ZipFile, rels_path: str, base_dir: str) -> dict[str, str]:
    """.rels を読み、rId → パッケージ内の正規化済みパス の辞書を返す。"""
    try:
        root = ET.fromstring(zf.read(rels_path))
    except (KeyError, ET.ParseError):
        return {}
    out: dict[str, str] = {}
    for rel in root:
        rid = rel.get("Id")
        target = rel.get("Target") or ""
        if not rid or rel.get("TargetMode") == "External":
            continue
        out[rid] = _normalize_part(target, base_dir)
    return out


def _normalize_part(target: str, base_dir: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    parts = [p for p in base_dir.split("/") if p]
    for segment in target.split("/"):
        if segment in ("", "."):
            continue
        if segment == "..":
            if parts:
                parts.pop()
        else:
            parts.append(segment)
    return "/".join(parts)


def _count_drawing_objects(root: ET.Element, found: dict[str, int]) -> None:
    """アンカー直下のオブジェクトを1件ずつ数える（グループ内は展開しない）。"""
    for anchor in root:
        if _local(anchor.tag) not in _ANCHOR_TAGS:
            continue
        for child in anchor:
            name = _local(child.tag)
            if name == "graphicFrame":
                label = "グラフ" if _is_chart_frame(child) else "その他のオブジェクト"
            elif name in _OBJECT_LABELS:
                label = _OBJECT_LABELS[name]
            else:
                continue
            found[label] = found.get(label, 0) + 1


def _is_chart_frame(frame: ET.Element) -> bool:
    for node in frame.iter():
        if _local(node.tag) == "graphicData":
            return "chart" in (node.get("uri") or "")
    return False


# ---------------------------------------------------------------------------
# シート → Markdown
# ---------------------------------------------------------------------------


def sheet_to_markdown(
    ws,
    *,
    formula_cells: set[tuple[int, int]] | None = None,
    drawings: dict[str, int] | None = None,
) -> tuple[str, bool]:
    """ワークシートをMarkdown本文へ変換する。

    戻り値は (本文, 数式のキャッシュ値が欠けていたか)。
    本文が空文字列の場合は「空シート」を意味する（spec.md 11.3）。
    """
    formula_cells = formula_cells or set()
    max_row, max_col = ws.max_row or 0, ws.max_column or 0
    if max_row < 1 or max_col < 1:
        return ("", False)

    # 結合セル: 左上以外は空欄にするための対応表と、罫線判定用の占有情報
    anchor_of: dict[tuple[int, int], tuple[int, int]] = {}
    merged_ranges = [
        (r.min_row, r.min_col, r.max_row, r.max_col) for r in ws.merged_cells.ranges
    ]
    for r1, c1, r2, c2 in merged_ranges:
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                anchor_of[(r, c)] = (r1, c1)

    grid: dict[tuple[int, int], object] = {}
    bordered: set[tuple[int, int]] = set()
    has_value = False
    missing_formula = False

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            key = (cell.row, cell.column)
            grid[key] = cell
            if _has_border(cell):
                bordered.add(key)
            if cell.value is not None:
                has_value = True
            elif key in formula_cells:
                # 数式なのにキャッシュ値が無い（spec.md 11.5）
                missing_formula = True

    # 結合セルは範囲全体を1つの占有セルとして扱う（spec.md 11.4）
    for r1, c1, r2, c2 in merged_ranges:
        cells = [(r, c) for r in range(r1, r2 + 1) for c in range(c1, c2 + 1)]
        if any(key in bordered for key in cells):
            bordered.update(cells)

    tables = detect_tables(bordered)
    if not has_value and not tables and not drawings:
        return ("", False)

    def text_at(r: int, c: int, *, in_table: bool) -> str:
        anchor = anchor_of.get((r, c))
        if anchor is not None and anchor != (r, c):
            return ""  # 結合セルは左上にのみ値を出す
        cell = grid.get((r, c))
        return cell_text(cell, in_table=in_table) if cell is not None else ""

    covered_rows: dict[int, list[Rect]] = {}
    for rect in tables:
        for r in range(rect[0], rect[2] + 1):
            covered_rows.setdefault(r, []).append(rect)

    blocks: list[str] = []
    paragraph: list[str] = []

    def flush_paragraph() -> None:
        if paragraph:
            # Markdownは単なる改行を空白として扱うため、そのまま連結すると
            # 複数行が1行に潰れてしまう。11.5のセル内改行と同じく <br> を使い、
            # Excelの行の区切りを保ったまま1つの段落にまとめる。
            blocks.append("<br>\n".join(paragraph))
            paragraph.clear()

    for r in range(1, max_row + 1):
        rects = covered_rows.get(r)
        if rects is not None:
            flush_paragraph()
            # このrowで始まるテーブルのみ、左の列から順に出力する
            for rect in sorted(rects, key=lambda x: x[1]):
                if rect[0] == r:
                    blocks.append(_render_table(rect, text_at))
            continue
        # 非テーブル行: 値の入ったセルを半角スペースで連結する（spec.md 11.6）
        values = [
            t for c in range(1, max_col + 1)
            if (t := text_at(r, c, in_table=False))
        ]
        if values:
            paragraph.append(" ".join(values))
        else:
            flush_paragraph()  # 値の無い行は段落の区切り
    flush_paragraph()

    if drawings:
        detail = "、".join(f"{label} {n}" for label, n in sorted(drawings.items()))
        blocks.append(
            "> **注記**: このシートには変換対象外の描画オブジェクト"
            f"（{detail}）が含まれています。"
        )

    return ("\n\n".join(blocks), missing_formula)


def _render_table(rect: Rect, text_at: Callable[..., str]) -> str:
    r1, c1, r2, c2 = rect
    width = c2 - c1 + 1
    lines = []
    for r in range(r1, r2 + 1):
        cells = [text_at(r, c, in_table=True) for c in range(c1, c2 + 1)]
        lines.append("| " + " | ".join(cells) + " |")
        if r == r1:
            # 先頭行を常にヘッダー行として扱う（spec.md 11.4）
            lines.append("|" + "---|" * width)
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# ブック全体の変換
# ---------------------------------------------------------------------------


def _collect_formula_cells(path: Path) -> dict[str, set[tuple[int, int]]]:
    """シートごとの数式セル座標を集める（キャッシュ値欠落の検出用）。"""
    out: dict[str, set[tuple[int, int]]] = {}
    try:
        wb = load_workbook(path, data_only=False, read_only=True)
    except Exception:
        return out
    try:
        for ws in wb.worksheets:
            coords: set[tuple[int, int]] = set()
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        coords.add((cell.row, cell.column))
            out[ws.title] = coords
    finally:
        wb.close()
    return out


def convert_workbook(
    xlsx_path: Path,
    out_dir: Path,
    progress: Callable[[int, int, str], None] | None = None,
) -> ConversionResult:
    """Excelブックをシート単位のMarkdownへ変換して out_dir に書き出す。

    シート単位で失敗しても中断せず、結果をまとめて返す（spec.md 11.7）。
    """
    xlsx_path = Path(xlsx_path)
    out_dir = Path(out_dir)
    result = ConversionResult(out_dir=out_dir)

    drawings = scan_drawing_objects(xlsx_path)
    formula_cells = _collect_formula_cells(xlsx_path)
    wb = load_workbook(xlsx_path, data_only=True)
    try:
        sheets = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
        file_names = unique_file_names([ws.title for ws in sheets])
        out_dir.mkdir(parents=True, exist_ok=True)

        for index, (ws, stem) in enumerate(zip(sheets, file_names), start=1):
            if progress is not None:
                progress(index, len(sheets), ws.title)
            try:
                body, missing = sheet_to_markdown(
                    ws,
                    formula_cells=formula_cells.get(ws.title, set()),
                    drawings=drawings.get(ws.title),
                )
                if not body:
                    result.sheets.append(
                        SheetResult(ws.title, ok=True, path=None, skipped="empty")
                    )
                    continue
                target = out_dir / f"{stem}.md"
                content = f"# {ws.title}\n\n{body}\n"
                target.write_text(content, encoding="utf-8", newline="\n")
                result.sheets.append(SheetResult(ws.title, ok=True, path=target))
                if missing:
                    result.missing_formula_sheets.append(ws.title)
            except Exception as e:  # 1シートの失敗で全体を止めない
                result.sheets.append(
                    SheetResult(ws.title, ok=False, error=f"{type(e).__name__}: {e}")
                )
    finally:
        wb.close()
    return result
