"""Excel → Markdown 変換（spec.md 11章）の検証。Qt非依存。

1. ファイル名のサニタイズと重複回避（11.2）
2. 罫線ベースのテーブル判定（11.4）
3. 表示形式・文字装飾・リンク・改行・エスケープ（11.5）
4. 非テーブル領域の段落化（11.6）
5. 出力順序・非表示/空シート・数式のキャッシュ値欠落（11.3 / 11.5 / 11.7）
6. 描画オブジェクトの検出と注記（11.8）

実行: .venv/bin/python tests/test_excel_import.py
"""
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Border, Font, Side
from openpyxl.worksheet.hyperlink import Hyperlink

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from markdown_editor.excel_import import (  # noqa: E402
    convert_workbook,
    decorate,
    detect_tables,
    format_value,
    sanitize_sheet_name,
    scan_drawing_objects,
    unique_file_names,
)

THIN = Side(style="thin")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
UNDERLINE = Border(bottom=THIN)

tmpdir = Path(tempfile.mkdtemp(prefix="md_excel_test_"))

# ---- 1. ファイル名のサニタイズ（spec.md 11.2） ----
assert sanitize_sheet_name("売上") == "売上"
assert sanitize_sheet_name('A<B>C"D|E') == "A_B_C_D_E"
assert sanitize_sheet_name("  前後空白  ") == "前後空白"
assert sanitize_sheet_name("末尾ピリオド...") == "末尾ピリオド"
assert sanitize_sheet_name("con") == "con_", sanitize_sheet_name("con")
assert sanitize_sheet_name("LPT1") == "LPT1_"
assert sanitize_sheet_name("   ") == "sheet"
assert sanitize_sheet_name("...") == "sheet"
print("SANITIZE OK")

assert unique_file_names(["A", "A", "A"]) == ["A", "A-2", "A-3"]
# 正規化後に衝突するケース（`<` と `>` がどちらも `_` になる）
assert unique_file_names(["A<B", "A>B"]) == ["A_B", "A_B-2"]
# ファイルシステムの大文字小文字非区別に合わせる
assert unique_file_names(["Sheet", "sheet"]) == ["Sheet", "sheet-2"]
print("UNIQUE NAMES OK")

# ---- 2. テーブル判定（spec.md 11.4） ----
# 3x3のべた塗り罫線 → 1つのテーブル
box = {(r, c) for r in range(1, 4) for c in range(1, 4)}
assert detect_tables(box) == [(1, 1, 3, 3)], detect_tables(box)

# 見出しの下線だけ（1行×3列） → 最小サイズ未満なのでテーブルにしない
assert detect_tables({(1, 1), (1, 2), (1, 3)}) == []
# 縦1列だけ → 同じくテーブルにしない
assert detect_tables({(1, 1), (2, 1), (3, 1)}) == []

# 離れた2つの領域 → 2つのテーブル
two = {(r, c) for r in range(1, 3) for c in range(1, 3)}
two |= {(r, c) for r in range(5, 7) for c in range(1, 3)}
assert detect_tables(two) == [(1, 1, 2, 2), (5, 1, 6, 2)], detect_tables(two)

# 連結していなくても外接矩形が交差する候補どうしは1つに統合される。
# L字（左下）と右上のブロックは接していないが、矩形は行1-2×列3で重なる。
l_shape = {(1, 1), (2, 1), (3, 1), (3, 2), (3, 3)}  # 外接矩形 (1,1,3,3)
block = {(1, 3), (1, 4), (2, 4)}  # 外接矩形 (1,3,2,4)
assert detect_tables(l_shape) == [(1, 1, 3, 3)]
assert detect_tables(block) == [(1, 3, 2, 4)]
assert detect_tables(l_shape | block) == [(1, 1, 3, 4)], detect_tables(l_shape | block)

# 矩形内の穴（罫線の無いセル）もテーブルの一部として扱う
ring = box - {(2, 2)}
assert detect_tables(ring) == [(1, 1, 3, 3)], detect_tables(ring)
print("TABLE DETECTION OK")

# ---- 3. 表示形式（spec.md 11.5） ----
cases = [
    (1234.5, "General", "1234.5"),
    (1234.0, "General", "1234"),
    # Excelは四捨五入（Pythonの既定である偶数丸めとは異なる）
    (1234.5, "#,##0", "1,235"),
    (2.5, "0", "3"),
    (1234.567, "#,##0.00", "1,234.57"),
    (0.1234, "0.0%", "12.3%"),
    (0.5, "0%", "50%"),
    (1234, '"¥"#,##0', "¥1,234"),
    (-1234, "#,##0;(#,##0)", "(1,234)"),
    (-1234, "#,##0", "-1,234"),
    (0, "#,##0;-#,##0;\"-\"", "-"),
    (7, "000", "007"),
    (datetime(2026, 8, 4), "yyyy/mm/dd", "2026/08/04"),
    (datetime(2026, 8, 4), 'yyyy"年"m"月"d"日"', "2026年8月4日"),
    (datetime(2026, 8, 4), "yy/m/d", "26/8/4"),
    (datetime(2026, 8, 4, 13, 5, 9), "h:mm:ss", "13:05:09"),
    (datetime(2026, 8, 4, 13, 5, 9), "h:mm AM/PM", "1:05 PM"),
    (datetime(2026, 8, 4), "mmm d, yyyy", "Aug 4, 2026"),
    (datetime(2026, 8, 4), "[$-409]dddd", "Tuesday"),
    ("そのまま", "General", "そのまま"),
    (None, "General", ""),
    (True, "General", "TRUE"),
]
for value, fmt, expected in cases:
    got = format_value(value, fmt)
    assert got == expected, f"format_value({value!r}, {fmt!r}) = {got!r} != {expected!r}"
print("NUMBER FORMAT OK")

assert decorate("値", bold=True, italic=False, strike=False) == "**値**"
assert decorate("値", bold=True, italic=True, strike=True) == "~~***値***~~"
assert decorate("値", bold=False, italic=False, strike=False) == "値"
# 前後の空白は装飾の外側に出す（Markdownの強調は空白に隣接すると成立しない）
assert decorate(" 値 ", bold=True, italic=False, strike=False) == " **値** "
print("DECORATION OK")

# ---- 4-6. 実ファイルを介した変換 ----
wb = Workbook()

ws = wb.active
ws.title = "売上"
# 見出し（下線のみ・1行）→ 段落として出力される
ws["A1"] = "2026年 上期 売上レポート"
ws["A1"].font = Font(bold=True)
ws["A1"].border = UNDERLINE
ws["A3"] = "単位: 千円"
# 罫線で囲まれた3x3のテーブル
rows = [["商品", "数量", "金額"], ["りんご|青", 10, 1234.5], ["みかん", 20, 987.0]]
for i, row in enumerate(rows):
    for j, value in enumerate(row):
        cell = ws.cell(row=5 + i, column=1 + j, value=value)
        cell.border = BOX
        if i == 0:
            cell.font = Font(bold=True)
ws["C6"].number_format = "#,##0"
ws["C7"].number_format = "#,##0"
# セル内改行 / パイプ / リンク / 装飾
ws["A9"] = "備考|注意\n2行目"
ws["A10"] = "公式サイト"
ws["A10"].hyperlink = Hyperlink(ref="A10", target="https://example.com/")
ws["A11"] = "取り消し"
ws["A11"].font = Font(strike=True)

ws2 = wb.create_sheet("結合と数式")
ws2["A1"] = "見出し"
ws2.merge_cells("A1:B1")
for coord in ("A1", "B1", "A2", "B2"):
    ws2[coord].border = BOX
ws2["A2"] = 1
ws2["B2"] = "=A2*2"  # data_only では計算結果が無いためキャッシュ値欠落になる

ws3 = wb.create_sheet("非表示")
ws3["A1"] = "出力されない"
ws3.sheet_state = "hidden"

wb.create_sheet("空シート")

ws5 = wb.create_sheet("グラフあり")
ws5["A1"] = "月"
ws5["B1"] = "売上"
for i, (m, v) in enumerate([("1月", 10), ("2月", 20), ("3月", 30)], start=2):
    ws5.cell(row=i, column=1, value=m)
    ws5.cell(row=i, column=2, value=v)
chart = BarChart()
chart.add_data(Reference(ws5, min_col=2, min_row=1, max_row=4), titles_from_data=True)
ws5.add_chart(chart, "D2")

ws6 = wb.create_sheet("見積書<案>")  # ファイル名として不正な文字を含む
ws6["A1"] = "山かっこ入り"

xlsx = tmpdir / "サンプル帳票.xlsx"
wb.save(xlsx)

# 描画オブジェクトの検出（zip直読み）
found = scan_drawing_objects(xlsx)
assert found == {"グラフあり": {"グラフ": 1}}, found
print("DRAWING SCAN OK")

out_dir = tmpdir / "サンプル帳票"
result = convert_workbook(xlsx, out_dir)

names = sorted(p.name for p in result.written)
assert names == ["グラフあり.md", "売上.md", "結合と数式.md", "見積書_案_.md"], names
assert not (out_dir / "非表示.md").exists(), "非表示シートは出力しない"
assert not (out_dir / "空シート.md").exists(), "空シートは出力しない"
assert [s.sheet_name for s in result.sheets if s.skipped == "empty"] == ["空シート"]
assert result.failed == [], result.failed
print("SHEET SELECTION OK")

body = (out_dir / "売上.md").read_text(encoding="utf-8")
expected = """\
# 売上

**2026年 上期 売上レポート**

単位: 千円

| **商品** | **数量** | **金額** |
|---|---|---|
| りんご\\|青 | 10 | 1,235 |
| みかん | 20 | 987 |

備考|注意<br>2行目<br>
[公式サイト](https://example.com/)<br>
~~取り消し~~
"""
assert body == expected, f"---got---\n{body}\n---want---\n{expected}"
print("SHEET BODY OK")

merged = (out_dir / "結合と数式.md").read_text(encoding="utf-8")
# 結合セルは左上のみ値を持ち、残りは空欄になる
assert merged == "# 結合と数式\n\n| 見出し |  |\n|---|---|\n| 1 |  |\n", repr(merged)
assert result.missing_formula_sheets == ["結合と数式"], result.missing_formula_sheets
print("MERGED CELL AND FORMULA OK")

chart_md = (out_dir / "グラフあり.md").read_text(encoding="utf-8")
assert "> **注記**" in chart_md and "グラフ 1" in chart_md, chart_md
print("DRAWING NOTE OK")

# 進捗コールバックは可視シートのみを対象に 1..n で呼ばれる
calls: list[tuple[int, int, str]] = []
convert_workbook(xlsx, out_dir, progress=lambda i, n, name: calls.append((i, n, name)))
assert [c[0] for c in calls] == [1, 2, 3, 4, 5], calls
assert all(c[1] == 5 for c in calls), calls
assert "非表示" not in [c[2] for c in calls], calls
print("PROGRESS OK")

# 2回目の変換は同名mdを上書きし、無関係なファイルは残す（spec.md 11.2）
keep = out_dir / "手書きメモ.md"
keep.write_text("残す", encoding="utf-8")
convert_workbook(xlsx, out_dir)
assert keep.read_text(encoding="utf-8") == "残す", "無関係なファイルを消してはいけない"
print("OVERWRITE KEEPS OTHERS OK")

shutil.rmtree(tmpdir, ignore_errors=True)
print("ALL EXCEL IMPORT TESTS PASSED")
