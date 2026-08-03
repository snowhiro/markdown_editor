"""Excel取り込みのアプリ側の挙動（spec.md 11.1 / 11.2 / 11.7）のオフスクリーン検証。

1. ファイルメニューに「Excelから変換...」がある
2. .xls を指定するとエラーになり、何も出力されない
3. 変換するとExcelと同階層に「ファイル名」フォルダができ、シート単位でmdが出る
4. 変換後、ツリーのルートが出力先になり先頭シートのmdが開かれる
5. 出力先が既存の場合は確認ダイアログを経由し、キャンセルすると何も書かない
6. 未保存の変更があると保存確認を経由し、キャンセルするとファイルを開かない
   （ツリーの表示は更新される）

実行: QT_QPA_PLATFORM=offscreen .venv/bin/python tests/test_excel_menu.py
"""
import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Side
from PySide6.QtWidgets import QApplication, QMessageBox
from markdown_editor.main import MainWindow

app = QApplication(sys.argv)

THIN = Side(style="thin")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

tmpdir = Path(tempfile.mkdtemp(prefix="md_excel_menu_"))
xlsx = tmpdir / "予算管理.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "概要"
ws["A1"] = "2026年度 予算"
ws2 = wb.create_sheet("明細")
for i, row in enumerate([["科目", "金額"], ["交通費", 1000], ["消耗品", 2000]]):
    for j, value in enumerate(row):
        cell = ws2.cell(row=1 + i, column=1 + j, value=value)
        cell.border = BOX
wb.save(xlsx)

w = MainWindow()
w.show()

# ---- 1. メニュー項目 ----
menu_bar = w.menuBar()
file_action = menu_bar.actions()[0]  # 参照を保持しないとQMenuが破棄される
file_menu = file_action.menu()
labels = [a.text() for a in file_menu.actions()]
assert "Excelから変換..." in labels, labels
assert w.import_excel_action.isEnabled()
print("MENU ITEM OK")

# ダイアログはオフスクリーンでは応答できないため、応答を差し替える
answers: list = []
shown: list[tuple[str, str]] = []
_orig_critical = QMessageBox.critical
_orig_warning = QMessageBox.warning
_orig_information = QMessageBox.information
_orig_question = QMessageBox.question


def _record(kind):
    def handler(parent, title, text, *a, **k):
        shown.append((kind, f"{title}\n{text}"))
        return answers.pop(0) if answers else QMessageBox.StandardButton.Ok

    return staticmethod(handler)


QMessageBox.critical = _record("critical")
QMessageBox.warning = _record("warning")
QMessageBox.information = _record("information")
QMessageBox.question = _record("question")

# ---- 2. 非対応形式 ----
xls = tmpdir / "旧形式.xls"
xls.write_bytes(b"dummy")
shown.clear()
w.import_excel(xls)
assert shown and shown[0][0] == "critical" and "xlsx" in shown[0][1], shown
assert not (tmpdir / "旧形式").exists(), "非対応形式で出力先を作ってはいけない"
print("XLS REJECTED OK")

# ---- 3-4. 変換とツリー/文書の反映 ----
shown.clear()
w.import_excel(xlsx)
out_dir = tmpdir / "予算管理"
assert out_dir.is_dir(), out_dir
assert sorted(p.name for p in out_dir.glob("*.md")) == ["明細.md", "概要.md"]
assert shown and shown[-1][0] == "information", shown
assert "成功 2件 / 失敗 0件" in shown[-1][1], shown[-1][1]
print("CONVERT OK")

assert w.tree_root == out_dir.resolve(), w.tree_root
# 先頭シート（ワークブック順）のmdが開かれる
assert w.current_path == (out_dir / "概要.md").resolve(), w.current_path
assert w.current_content.startswith("# 概要"), w.current_content
print("TREE ROOT AND FIRST SHEET OK")

detail = (out_dir / "明細.md").read_text(encoding="utf-8")
assert "| 科目 | 金額 |" in detail and "| 交通費 | 1000 |" in detail, detail
print("TABLE OUTPUT OK")

# ---- 5. 既存フォルダの確認ダイアログでキャンセル ----
(out_dir / "概要.md").write_text("手で書き換えた", encoding="utf-8")
shown.clear()
answers.append(QMessageBox.StandardButton.Cancel)
w.import_excel(xlsx)
assert shown and shown[0][0] == "question", shown
assert (out_dir / "概要.md").read_text(encoding="utf-8") == "手で書き換えた", "キャンセル時は書き換えない"
print("OVERWRITE CANCEL OK")

# OKなら上書きする
shown.clear()
answers.append(QMessageBox.StandardButton.Ok)  # 上書き確認
w.import_excel(xlsx)
assert (out_dir / "概要.md").read_text(encoding="utf-8").startswith("# 概要")
print("OVERWRITE OK")

# ---- 6. 未保存の変更があるとき、保存確認をキャンセルすると開かない ----
w.on_content_changed(w.current_content + "\n編集した")
assert w.dirty
w._set_tree_root(tmpdir)  # ルートをずらしておき、更新されることを確かめる
shown.clear()
answers.append(QMessageBox.StandardButton.Ok)  # 上書き確認
answers.append(QMessageBox.StandardButton.Ok)  # 変換結果の通知
answers.append(QMessageBox.StandardButton.Cancel)  # 未保存の確認 → キャンセル
before = w.current_path
w.import_excel(xlsx)
assert w.current_path == before, "キャンセル時はファイルを開かない"
assert w.tree_root == out_dir.resolve(), "キャンセルしてもツリーの表示は更新する"
print("DISCARD CANCEL OK")

QMessageBox.critical = _orig_critical
QMessageBox.warning = _orig_warning
QMessageBox.information = _orig_information
QMessageBox.question = _orig_question

shutil.rmtree(tmpdir, ignore_errors=True)
print("ALL EXCEL MENU TESTS PASSED")
